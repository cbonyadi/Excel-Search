# -*- coding: utf-8 -*-
"""
Excel Lookup Python Program

Cyrus Jian Bonyadi
"""
import openpyxl as xl
import argparse
import difflib as dl

#open the file into a reader.
def open_file(name):
    """
    The open_file() method is designed to be used by the controller method
    to open a file and fix any problems with the file name provided by the
    user.
    
    Args:
        A file name as a string.
    
    Returns:
        A workbook file
    """
    file = None;
    
    while file == None:
        
        if name == '':
            #If we don't have a file name, let's get one.
            name = input("Please enter a file name.\n");
        
        #If we have a file extension, let's make sure it's the right one.
        elif name.find('.') + 1:
            print(name[name.find('.')::]);
            
            #If our file extension is correct, we accept it.
            if name[name.find('.')::] == ".xlsx" or name[name.find('.')::] == ".xlsm" or name[name.find('.')::] == ".xltx" or name[name.find('.')::] == ".xltm":
                try:
                    file = xl.load_workbook(name);
                    break;
                except IOError as e:
                    print("I/O error({0}): {1}".format(e.errno, e.strerror) + ": " + name);
                    name = '';
            
            #If it's incorrect, let's let the user know and start over.
            else:
                print("This program only works with .xltx, .xltm, .xlsx, and .xlsm files.\n");
                name = '';
                
        
        #If we don't have a file extension, let's make an assumption.
        else:
            name += ".xl";
            
            #Let's try to open the file.
            try:
                file = xl.load_workbook(name + "sx");
                break;
            except IOError as e:
                try:
                    file = xl.load_workbook(name + "sm");
                    break;
                except IOError as f:
                    try:
                        file = xl.load_workbook(name + "tx");
                        break;
                    except IOError as g:
                        try:
                            file = xl.load_workbook(name + "tm");
                            break;
                        except IOError as h:
                            print("I/O error({0}): {1}".format(h.errno, h.strerror));
                            name = '';
                        
    
    return file;
    

def get_sheet(workbook, name):
    """
    get_sheet grabs the name of the sheet in the workbook.
    
    Args:
        A workbook.
        A sheet name.
    Returns:
        A worksheet.
    """
    
    
    sheet = None;
    
    while sheet == None:
        #see if we even have a name
        if name == '':
            name = input("Please enter a sheet name.\n");
        #check if the name is found.
        elif name in workbook.sheetnames:
            sheet = workbook[name];
        #if it isn't found, let's reset and start over.
        else:
            print("Invalid sheet name: " + name +".\n");
            name = '';
    
    return sheet;
            
    

#if the column isn't valid, we should prompt the user for a valid column.
def get_columns(sheet):
    """
    This method is designed to go unseen by the user, creating a dictionary
    of column names alongside their alphabet letter.
    
    Args:
        An openpyxl sheet.
        
    Returns:
        A dictionary of column names and alphabet letters.
    """
    column_headers = {};
    
    #for all of the columns, let's make a dictionary of the column headers.
    for i in range(sheet.min_column - 1, sheet.max_column):
        column_headers[list(sheet.rows)[0][i].value]= i;
    
    return column_headers;

#lookup our index.
def lookup_indeces(sheet, presets, column_headers):
    """
    The lookup_index method finds the row of indeces that contain the 
    search term.
    
    Args:
        An excel sheet.
        Presets from the args.
        The column headers dictionary.
        
    Returns:
        An array of matching rows.
    """    
    indeces = [];
    column = presets.lookup_column;
    
    while not indeces or column == '':
        #grab a list of the column
        if column == '':
            column = input("Please enter a column name or number, where A=0.\n");
        elif type(column) is not int and not column.isdigit() and column in column_headers.keys():
            column = column_headers[column];
        elif type(column) is not int and (not column.isdigit() or int(column) not in range(sheet.min_column-1, sheet.max_column)):
            print("Invalid column.\n");
            column = '';
        else:
            column = int(column);
            cur_column = list(sheet.columns)[column];
            break;
            
    term = presets.lookup;
    
    while not indeces or term == '':
        if term == '':
            term = input("Please enter a search term.\n");
        else:
            if presets.not_exact:
                for cell in cur_column[1::]:
                    #find all the indeces of the column that match our search keyword
                    if dl.SequenceMatcher(None, term, cell.value).ratio() > 0:
                        #put these secondary with a percent match as the first item to be sorted
                        indeces.append(tuple((dl.SequenceMatcher(None, term, cell.value).ratio(), cell.row)));
                        
            else:
                for cell in cur_column[1::]:
                    #find all the indeces of the column that contain our search keyword
                    if term.upper() in cell.value.upper():
                        indeces.append(tuple((dl.SequenceMatcher(None, term, cell.value).ratio(), cell.row)));
                    
                    #put these secondary with a percent containing as the first item to be sorted
    
    return column, sorted(indeces, key=lambda tup: tup[0], reverse=True);
    
def print_indeces(sheet, indeces, presets):
    """
    return_indeces grabs indeces of a file and returns a most sized list of 
    the rows with the specified columns.
    
    Args:
        The excel sheet.
        An indeces list.
        Presets from the args.
        
    Returns:
        Nothing, it just prints.
    """
    if not indeces:
        print("No results found.\n");
        
    #if desired columns is empty, let's print up to most.
    else:
        for cell in list(sheet.rows)[0]:
            print(cell.value, end='\t'); #print each column header on one line
        print(); #newline
        
        if presets.most >= 0 and presets.most <= len(indeces):
            for match in indeces[0:most]:
                for cell in list(sheet.rows)[match[1]-1]:
                    print(cell.value, end='\t'); #print each cell on one line
                print(); #newline
        else:
            for match in indeces:
                for cell in list(sheet.rows)[match[1]-1]:
                    print(cell.value, end='\t'); #print each cell on one line
                print(); #newline
    

def menu_controller(presets):
    """
    menu_controller takes all the presets specified and begins to function
    as a controller for the lookup.
    
    Args:
        preset variables.
    Returns:
        nothing.
    """
    #grab our file reference and its column headers
    
    print(presets);
    
    excel_file = open_file(presets.file);
    
    worksheet = get_sheet(excel_file, presets.sheet);
    
    column_headers = get_columns(worksheet);
    
    #keep our controller open.    
    while True:        
        #get the indeces matching our lookup term
        column, indeces = lookup_indeces(worksheet, presets, column_headers);
        
        print_indeces(worksheet, indeces, presets);
        
        #ask if we're done, or continue.
        
    

def argument_parser():
    """
    argument_parser is designed to handle all the defaults provided by the
    user, creating a dictionary of default values to be used in the 
    controller.
    
    Args:
        The input arguments.
        
    Returns:
        The arguments or their defaults.
    """
    
    parser = argparse.ArgumentParser(description="A python file used to "+
                     "lookup rows in an excel file.  If you intend on putting"
                     + " a term with a space in it, put quotes around it.");
    
    parser.add_argument("-f", "--file", nargs="?", default="", 
                help="The name of the excel file.");
    parser.add_argument("-s", "--sheet", nargs="?", default="Sheet1", 
                help="The sheet of the excel file.  Sheet1 by default.");
    parser.add_argument("-lc", "--lookup_column", nargs="?", default="", 
                help="Specify which column to look at.  Quotations " 
                +"denote name, as opposed to numberical index (A=0).");
    parser.add_argument("-l", "--lookup", nargs="?", default="", 
                help="Use followed by search term.");
    parser.add_argument("-ne", "--not_exact", action="store_true", 
                help="Add this flag if you're accepting non-exact matches.");
    parser.add_argument("-m", "--most", nargs="?", default=-1, type=int,
                help="Use to get most alike in lookup.  "
                + "All are displayed by default.");
                
    return parser.parse_args();

def main():
    """
    Our main should just process arguments and initiating the
    controller.
    """
    
    #check if we have arguments
    #if we have arguments, let's open the argument parser.
    arg_vars = argument_parser();
    
    #if we're done processing arguments, let's go to the controller.
    menu_controller(arg_vars);
    
if __name__== "__main__":
    main();